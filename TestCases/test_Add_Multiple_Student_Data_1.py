import openpyxl
import json

import requests


def test_Add_multiple_students():
    # API
    API_url = "http://thetestingworldapi.com/api/studentsDetails"

    # Parsing the Json from Extenal file
    file = open("C:/Users/Sunil.more/Desktop/PythonScrap/Add_Student.json", "r")
    Json_request = json.load(file)

    # Loading Excel and finding workbook, sheet and rows
    workbook = openpyxl.load_workbook("C:/Users/Sunil.more/Desktop/PythonScrap/data.xlsx")
    sheet = workbook['Sheet1']
    rows = sheet.max_row

    for i in range(2, rows + 1):
        cell_first_name = sheet.cell(row=i, column=1)
        cell_middle_name = sheet.cell(row=i, column=2)
        cell_last_name = sheet.cell(row=i, column=3)
        cell_dob = sheet.cell(row=i, column=4)

        Json_request['first_name'] = cell_first_name.value
        Json_request['middle_name'] = cell_middle_name.value
        Json_request['last_name'] = cell_last_name.value
        Json_request['date_of_birth'] = cell_dob.value

        response = requests.post(API_url, Json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
