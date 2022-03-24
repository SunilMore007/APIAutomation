import openpyxl
import json
import requests


class Comman:

    # Constructor to fetch comman statement
    def __init__(self, FileNamePath, SheetName):
        global workbook
        global sheet
        workbook = openpyxl.load_workbook(FileNamePath)  # C:/Users/Sunil.more/Desktop/PythonScrap/data.xlsx
        sheet = workbook[SheetName]  # Sheet 1

    def fetch_row_count(self):
        # Loading Excel and finding workbook, sheet and rows

        rows = sheet.max_row
        return rows

    def fetch_column_count(self):
        column = sheet.max_column
        return column

    def fetch_Key_names(self):
        c = sheet.max_column

        li = []

        for i in range(1, c + 1):
            cell = sheet.cell(row=1, column=i)
            li.insert(i - 1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, json_request, Keylist):
        c = sheet.max_column
        for i in range(1, c + 1):
            cell = sheet.cell(row=rowNumber, column=i)
            json_request[Keylist[i - 1]] = cell.value

        return json_request
